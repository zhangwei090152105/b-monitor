#!/usr/bin/env python3
"""
吉客云 B 单未完结监控 - GitHub Actions 适配版

整合了 API 拉取 → Excel 生成 → 邮件发送 → 企微推送 → 看板 HTML 五个阶段的完整流水线。
所有敏感信息从环境变量读取，无硬编码，可在 GitHub Actions 中独立运行。

看板使用 v2.0 精装修模板（ECharts图表、暗色/亮色主题、玻璃拟态、筛选搜索、分页表格）。

用法:
    # 完整流水线（使用当天日期）
    python main.py

    # 指定日期
    B_MONITOR_DATE=20260711 python main.py
"""

import hashlib
import json
import os
import smtplib
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置（全部从环境变量读取）
# ============================================================

JKY_APPKEY    = os.environ.get('JKY_APPKEY', '')
JKY_APPSECRET = os.environ.get('JKY_APPSECRET', '')
JKY_API_URL   = 'https://open.jackyun.com/open/openapi/do'

SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.qq.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '465'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
TO_ADDR   = os.environ.get('TO_ADDR', '')
CC_ADDR   = os.environ.get('CC_ADDR', '')

WECOM_WEBHOOKS = [
    url for url in [
        os.environ.get('WECOM_WEBHOOK_1', ''),
        os.environ.get('WECOM_WEBHOOK_2', ''),
    ] if url and 'YOUR_KEY' not in url
]

MIAODA_URL = os.environ.get('MIAODA_URL', '')

# 数据日期
B_MONITOR_DATE = os.environ.get('B_MONITOR_DATE', '') or date.today().strftime('%Y%m%d')

# 输出目录（相对于脚本所在目录）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, 'output')
os.makedirs(OUT_DIR, exist_ok=True)

# 模板文件路径（v2.0 精装修模板前段/后段，不含 DATA 数据）
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, 'templates')
TEMPLATE_BEFORE_PATH = os.environ.get(
    'TEMPLATE_BEFORE_PATH',
    os.path.join(TEMPLATE_DIR, 'template_before.txt')
)
TEMPLATE_AFTER_PATH = os.environ.get(
    'TEMPLATE_AFTER_PATH',
    os.path.join(TEMPLATE_DIR, 'template_after.txt')
)

# ============================================================
# 阶段 1: API 调用层
# ============================================================

STOCKOUT_COLS = (
    'outNo,outWarehouseName,outWarehouseCode,outType,outTypeName,outStatus,'
    'applyDate,gmtCreate,memo,outReason,field1,'
    'goodsName,skuCount,outedCount,unOutedCount,totalSkuCount,'
    'totalOutedCount,totalUnoutedCount,'
    'skuBarcode,isCertified,relDataId,status'
)

STOCKIN_COLS = (
    'inNo,inWarehouseName,inWarehouseCode,inType,inTypeName,inStatus,'
    'applyDate,gmtCreate,memo,inReason,field1,'
    'goodsName,skuCount,innerCount,uninnerCount,totalSkuCount,'
    'totalInnerCount,totalUninnerCount,'
    'skuBarcode,isCertified,relDataId,brand,brandName,status,'
    'logisticNo,logisticName'
)


def md5_sign(params):
    """签名：APPSECRET + 按字母排序的(k+v) + APPSECRET，md5 小写"""
    items = {k: v for k, v in params.items() if k not in ('sign', 'contextid')}
    sorted_str = ''.join(f'{k}{v}' for k, v in sorted(items.items()))
    signed = JKY_APPSECRET + sorted_str + JKY_APPSECRET
    return hashlib.md5(signed.lower().encode('utf-8')).hexdigest()


def call_api(method, bizcontent, max_retry=3):
    """通用 API 调用（带重试 3 次，间隔 2s）"""
    for attempt in range(max_retry):
        try:
            ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            params = {
                'method': method, 'appkey': JKY_APPKEY, 'version': 'v1.0',
                'contenttype': 'json', 'timestamp': ts,
                'bizcontent': bizcontent
            }
            params['sign'] = md5_sign(params)
            data = urllib.parse.urlencode(params).encode('utf-8')
            req = urllib.request.Request(JKY_API_URL, data=data,
                headers={'Content-Type': 'application/x-www-form-urlencoded'},
                method='POST')
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            print(f'  [API] 重试 {attempt+1}/{max_retry}: {e}')
            time.sleep(2)
    return {'code': -1, 'msg': f'调用失败 {max_retry} 次'}


def _month_ranges():
    """生成 2026-01 到当前月的逐月日期范围"""
    ranges = []
    now = datetime.now()
    for y in range(2026, now.year + 1):
        end_m = 12 if y < now.year else now.month
        for m in range(1, end_m + 1):
            from_d = f'{y}-{m:02d}-01'
            if m in [1, 3, 5, 7, 8, 10, 12]:
                end_day = 31
            elif m in [4, 6, 9, 11]:
                end_day = 30
            else:
                end_day = 29 if y % 4 == 0 and (y % 100 != 0 or y % 400 == 0) else 28
            to_d = f'{y}-{m:02d}-{end_day}'
            ranges.append((from_d, to_d))
    return ranges


def fetch_stockout(page_size=100, max_total_seconds=600):
    """拉全部出库申请单（未完结 outStatus in 1,2），过滤 status=3"""
    all_data = []
    seen = set()
    start_ts = time.time()
    months = _month_ranges()
    total_pages = 0

    for from_d, to_d in months:
        for status in [1, 2]:
            page = 0
            while True:
                elapsed = time.time() - start_ts
                if elapsed > max_total_seconds:
                    print(f'  [出库] 超时 {elapsed:.0f}s，停止')
                    break
                bizcontent = {
                    'pageIndex': page, 'pageSize': page_size,
                    'outStatus': str(status),
                    'applyDateFrom': from_d, 'applyDateTo': to_d,
                    'cols': STOCKOUT_COLS,
                }
                res = call_api('erp.stockout.get.v2', json.dumps(bizcontent, separators=(',', ':')))
                if res.get('code') != 200:
                    print(f'  [出库 {from_d}~{to_d} status={status}] API错误: {res.get("msg")}')
                    break
                data = res.get('result', {}).get('data')
                if not data or not isinstance(data, list):
                    if not data:
                        break
                    data = [data]
                if len(data) == 0:
                    break
                for item in data:
                    no = item.get('outNo')
                    if not no or no in seen:
                        continue
                    if str(item.get('status')) == '3':
                        continue
                    seen.add(no)
                    all_data.append(item)
                page += 1
                total_pages += 1
                if len(data) < page_size:
                    break
                if page > 500:
                    break
            if elapsed > max_total_seconds:
                break
        if elapsed > max_total_seconds:
            break
    print(f'  [出库] 总计 {len(all_data)} 条, {total_pages} 页 (去重+过滤后)')
    return all_data


def fetch_stockin(page_size=100, max_total_seconds=600):
    """拉全部入库申请单（未完结 inStatus in 1,2），过滤 status=3 + inType=105"""
    all_data = []
    seen = set()
    start_ts = time.time()
    months = _month_ranges()
    total_pages = 0

    for from_d, to_d in months:
        for status in [1, 2]:
            page = 0
            while True:
                elapsed = time.time() - start_ts
                if elapsed > max_total_seconds:
                    print(f'  [入库] 超时 {elapsed:.0f}s，停止')
                    break
                bizcontent = {
                    'pageIndex': page, 'pageSize': page_size,
                    'inStatus': str(status),
                    'applyDateFrom': from_d, 'applyDateTo': to_d,
                    'cols': STOCKIN_COLS,
                }
                res = call_api('erp.stockin.get.v2', json.dumps(bizcontent, separators=(',', ':')))
                if res.get('code') != 200:
                    print(f'  [入库 {from_d}~{to_d} status={status}] API错误: {res.get("msg")}')
                    break
                data = res.get('result', {}).get('data')
                if not data or not isinstance(data, list):
                    if not data:
                        break
                    data = [data]
                if len(data) == 0:
                    break
                for item in data:
                    no = item.get('inNo')
                    if not no or no in seen:
                        continue
                    if str(item.get('status')) == '3':
                        continue
                    if str(item.get('inType')) == '105':
                        continue
                    seen.add(no)
                    all_data.append(item)
                page += 1
                total_pages += 1
                if len(data) < page_size:
                    break
                if page > 500:
                    break
            if elapsed > max_total_seconds:
                break
        if elapsed > max_total_seconds:
            break
    print(f'  [入库] 总计 {len(all_data)} 条, {total_pages} 页 (去重+过滤后)')
    return all_data


def fetch_brand_map(sku_barcodes):
    """按 skuBarcode 批量反查 brandName"""
    barcode_to_brand = {}
    barcodes_list = list(sku_barcodes)
    for i in range(0, len(barcodes_list), 20):
        batch = barcodes_list[i:i+20]
        bizcontent = json.dumps({'skuBarcodes': ','.join(batch)}, separators=(',', ':'))
        res = call_api('erp.storage.goodslist', bizcontent)
        if res.get('code') != 200:
            print(f'  [货品反查] 批次 {i//20+1} 失败: {res.get("msg")}')
            continue
        data = res.get('result', {}).get('data', {})
        goods_list = data.get('goods') or []
        for g in goods_list:
            bc = g.get('skuBarcode', '')
            bn = g.get('brandName', '')
            if bc and bn:
                barcode_to_brand[bc] = bn
        print(f'  [货品反查] 批次 {i//20+1}: {len(batch)} 条 → 命中 {len(goods_list)}')
    return barcode_to_brand


def save_json(data, prefix):
    """落盘到 output/"""
    path = os.path.join(OUT_DIR, f'{prefix}_{B_MONITOR_DATE}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'  [保存] {path} ({len(data)} 条)')


# ============================================================
# 阶段 2: Excel 生成
# ============================================================

def get_status_name(status, biz_type='out'):
    if biz_type == 'out':
        return {0: '待递交', 1: '审核中', 2: '审核通过', 3: '关闭', 4: '作废'}.get(status, str(status))
    else:
        return {0: '待递交', 1: '待审核', 2: '已审核', 3: '已关闭', 10: '审核中'}.get(status, str(status))


def build_summary(data, group_keys, prefix='out'):
    groups = {}
    for x in data:
        key = tuple(x.get(k, '未知') for k in group_keys)
        if key not in groups:
            groups[key] = {'count': 0, 'total_qty': 0.0, 'total_done': 0.0, 'total_remain': 0.0}
        groups[key]['count'] += 1
        groups[key]['total_qty'] += float(x.get('totalSkuCount') or 0)
        groups[key]['total_done'] += float(
            x.get('totalOutedCount' if prefix == 'out' else 'totalInnerCount') or 0)
        groups[key]['total_remain'] += float(
            x.get('totalUnoutedCount' if prefix == 'out' else 'totalUninnerCount') or 0)
    return groups


def write_summary_sheet(ws, groups, label_prefix='出库'):
    if label_prefix == '出库':
        headers = ['业务类型', '品牌', '仓库', '单数', '申请数量', '未出库数', '出库率']
    else:
        headers = ['业务类型', '品牌', '仓库', '单数', '申请数量', '未入库数', '入库率']

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, (key, v) in enumerate(sorted(groups.items()), 2):
        done = v['total_done']
        qty = v['total_qty']
        rate = (done / qty * 100) if qty > 0 else 0
        cells = list(key) + [
            v['count'], int(v['total_qty']), int(v['total_remain']),
            f'{rate:.1f}%',
        ]
        for col_idx, val in enumerate(cells, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def write_detail_sheet(ws, data, biz_type='out'):
    if biz_type == 'out':
        headers = [
            '申请时间', '仓库名称', '出库类型名称', '出库申请单编号', '关联单据编号',
            '申请单备注', '出库原因', '店铺',
            '货品总数量(主表)', '已出库数量(主表)', '未出库数量(主表)',
            '货品品牌名称', '货品名称', '条形码',
            '数量', '已出库数量(明细)', '未出库数量(明细)', '是否正品'
        ]
    else:
        headers = [
            '申请时间', '仓库名称', '入库类型名称', '入库申请单号', '关联单号',
            '物流公司名称', '物流单号', '备注', '入库原因', '店铺',
            '货品总数量(主表)', '已入库数量(主表)', '未入库数量(主表)',
            '货品品牌名称', '货品名称', '条形码',
            '数量', '已入库数量(明细)', '未入库数量(明细)', '是否正品'
        ]

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, x in enumerate(data, 2):
        apply_date = ''
        if x.get('applyDate'):
            try:
                apply_date = datetime.fromtimestamp(x['applyDate'] / 1000).strftime('%Y-%m-%d %H:%M')
            except (OSError, ValueError, TypeError):
                pass

        if biz_type == 'out':
            row_data = [
                apply_date,
                x.get('outWarehouseName', ''),
                x.get('outTypeName', ''),
                x.get('outNo', ''),
                x.get('relDataId', ''),
                x.get('memo', ''),
                x.get('outReason', ''),
                x.get('field1', ''),
                int(x.get('totalSkuCount') or 0),
                int(x.get('totalOutedCount') or 0),
                int(x.get('totalUnoutedCount') or 0),
                x.get('brandName', ''),
                x.get('goodsName', ''),
                x.get('skuBarcode', ''),
                int(x.get('skuCount') or 0),
                int(x.get('outedCount') or 0),
                int(x.get('unOutedCount') or 0),
                '是' if x.get('isCertified') else '否',
            ]
        else:
            row_data = [
                apply_date,
                x.get('inWarehouseName', ''),
                x.get('inTypeName', ''),
                x.get('inNo', ''),
                x.get('relDataId', ''),
                x.get('logisticName', '') or '',
                x.get('logisticNo', '') or '',
                x.get('memo', ''),
                x.get('inReason', ''),
                x.get('field1', ''),
                int(x.get('totalSkuCount') or 0),
                int(x.get('totalInnerCount') or 0),
                int(x.get('totalUninnerCount') or 0),
                x.get('brandName', ''),
                x.get('goodsName', ''),
                x.get('skuBarcode', ''),
                int(x.get('skuCount') or 0),
                int(x.get('innerCount') or 0),
                int(x.get('uninnerCount') or 0),
                '是' if x.get('isCertified') else '否',
            ]

        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = 'A2'


def generate_excel(stockout, stockin):
    """生成 4 sheet Excel，返回文件路径"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. 入库申请汇总
    ws = wb.create_sheet('1.入库申请汇总')
    in_groups = build_summary(stockin, ['inTypeName', 'brandName', 'inWarehouseName'], prefix='in')
    write_summary_sheet(ws, in_groups, label_prefix='入库')

    # 2. 入库申请明细
    ws = wb.create_sheet('2.入库申请明细')
    write_detail_sheet(ws, stockin, biz_type='in')

    # 3. 出库申请汇总
    ws = wb.create_sheet('3.出库申请汇总')
    out_groups = build_summary(stockout, ['outTypeName', 'brandName', 'outWarehouseName'], prefix='out')
    write_summary_sheet(ws, out_groups, label_prefix='出库')

    # 4. 出库申请明细
    ws = wb.create_sheet('4.出库申请明细')
    write_detail_sheet(ws, stockout, biz_type='out')

    out_path = os.path.join(OUT_DIR, f'jkyun_b_monitor_{B_MONITOR_DATE}.xlsx')
    wb.save(out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f'✅ Excel 生成成功: {out_path} ({size_kb:.0f} KB)')
    return out_path


# ============================================================
# 阶段 3: 邮件发送
# ============================================================

def _safe_int(n):
    return n if n is None else int(n)


def send_email(excel_path, stockout, stockin):
    """发送 HTML 正文邮件 + Excel 附件"""
    # 构建统计摘要
    so_cnt = len(stockout)
    so_total = sum(float(x.get('totalSkuCount') or 0) for x in stockout)
    so_remain = sum(float(x.get('totalUnoutedCount') or 0) for x in stockout)
    so_rate = (so_total - so_remain) / so_total * 100 if so_total > 0 else 0
    so_brands = len(set(x.get('brandName', '') for x in stockout))

    si_cnt = len(stockin)
    si_total = sum(float(x.get('totalSkuCount') or 0) for x in stockin)
    si_remain = sum(float(x.get('totalUninnerCount') or 0) for x in stockin)
    si_rate = (si_total - si_remain) / si_total * 100 if si_total > 0 else 0
    si_brands = len(set(x.get('brandName', '') for x in stockin))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    html = f"""<html><body style="font-family: 'Microsoft YaHei', Arial, sans-serif; max-width: 680px;">

<h2>📦 吉客云 B 单未完结监控</h2>
<p style="color: #888; font-size: 13px;">数据日期: {B_MONITOR_DATE} · 生成时间: {now_str}</p>

<h3>📊 核对统计</h3>
<table style="border-collapse: collapse; width: 100%; text-align: center;">
<tr style="background: #1F4E78; color: white;">
  <th style="padding: 8px;">业务类型</th>
  <th style="padding: 8px;">单数</th>
  <th style="padding: 8px;">品牌数</th>
  <th style="padding: 8px;">申请数量</th>
  <th style="padding: 8px;">未完结数</th>
  <th style="padding: 8px;">完结率</th>
</tr>
<tr style="background: #FFF3E0;">
  <td style="padding: 8px; border: 1px solid #ddd;">📤 出库申请</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{so_cnt:,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{so_brands}</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{_safe_int(so_total):,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;"><b style="color: red;">{_safe_int(so_remain):,}</b></td>
  <td style="padding: 8px; border: 1px solid #ddd;">{so_rate:.1f}%</td>
</tr>
<tr style="background: #E3F2FD;">
  <td style="padding: 8px; border: 1px solid #ddd;">📥 入库申请</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{si_cnt:,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{si_brands}</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{_safe_int(si_total):,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;"><b style="color: red;">{_safe_int(si_remain):,}</b></td>
  <td style="padding: 8px; border: 1px solid #ddd;">{si_rate:.1f}%</td>
</tr>
<tr style="background: #F5F5F5; font-weight: bold;">
  <td style="padding: 8px; border: 1px solid #ddd;">合计</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{so_cnt + si_cnt:,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;">-</td>
  <td style="padding: 8px; border: 1px solid #ddd;">{_safe_int(so_total + si_total):,}</td>
  <td style="padding: 8px; border: 1px solid #ddd;"><b style="color: red;">{_safe_int(so_remain + si_remain):,}</b></td>
  <td style="padding: 8px; border: 1px solid #ddd;">-</td>
</tr>
</table>

<h3>📎 附件说明</h3>
<p>jkyun_b_monitor_{B_MONITOR_DATE}.xlsx — Excel 报告 (4 sheet: 出库汇总/明细、入库汇总/明细)</p>

<h3>🌐 数字看板</h3>
<p>🔗 <a href="{MIAODA_URL}">{MIAODA_URL}</a></p>

<hr style="border: none; border-top: 1px solid #ddd;">
<p style="color: #999; font-size: 12px;">
本邮件由 GitHub Actions 自动生成 · 每天 9:00 (北京时间) 自动发送<br>
数据来源：吉客云 Open API
</p>
</body></html>"""

    msg = MIMEMultipart()
    msg['Subject'] = f'【吉客云B单未完结监控】{B_MONITOR_DATE}'
    msg['From'] = SMTP_USER
    msg['To'] = TO_ADDR
    if CC_ADDR:
        msg['Cc'] = CC_ADDR

    msg.attach(MIMEText(html, 'html', 'utf-8'))

    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(excel_path)}"')
        msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)

    print(f'✅ 邮件发送成功 → {TO_ADDR}' + (f' (抄送 {CC_ADDR})' if CC_ADDR else ''))


# ============================================================
# 阶段 4: 企微推送
# ============================================================

def send_wecom(stockout, stockin):
    """企业微信 Markdown 推送（Top 5 摘要）"""
    so_cnt = len(stockout)
    so_remain = sum(float(x.get('totalUnoutedCount') or 0) for x in stockout)
    si_cnt = len(stockin)
    si_remain = sum(float(x.get('totalUninnerCount') or 0) for x in stockin)

    def top5(data, remain_field, type_name_field):
        agg = defaultdict(lambda: {'count': 0, 'remain': 0.0, 'type_name': ''})
        for x in data:
            brand = x.get('brandName', '未知')
            key = (brand, x.get(type_name_field, ''))
            if agg[key]['count'] == 0:
                agg[key]['type_name'] = x.get(type_name_field, '')
            agg[key]['count'] += 1
            agg[key]['remain'] += float(x.get(remain_field) or 0)
        sorted_items = sorted(agg.items(), key=lambda kv: kv[1]['remain'], reverse=True)[:5]
        return [(brand, d['type_name'], d['count'], d['remain']) for (brand, _), d in sorted_items]

    so_top5 = top5(stockout, 'totalUnoutedCount', 'outTypeName')
    si_top5 = top5(stockin, 'totalUninnerCount', 'inTypeName')

    parts = []
    parts.append(f'📢 吉客云B单未完结监控 · {B_MONITOR_DATE}')
    parts.append('\n\n\n')
    parts.append('📊 汇总\n')
    parts.append(f'　出库 {so_cnt:,} 单 / 待出 {int(so_remain):,} 件\n')
    parts.append(f'　入库 {si_cnt:,} 单 / 待入 {int(si_remain):,} 件')
    parts.append('\n\n\n')
    parts.append('🔴 出库 Top 5\n')
    for i, (brand, tp, n, q) in enumerate(so_top5, 1):
        parts.append(f'{i}. {brand}·{tp}　{n}单 / 待出 {int(q):,}件\n')
    parts.append('\n\n\n')
    parts.append('🔵 入库 Top 5\n')
    for i, (brand, tp, n, q) in enumerate(si_top5, 1):
        parts.append(f'{i}. {brand}·{tp}　{n}单 / 待入 {int(q):,}件\n')
    parts.append('\n\n\n')
    if MIAODA_URL:
        parts.append(f'[查看数字看板]({MIAODA_URL})')
        parts.append('\n\n')
    parts.append(f'🕐 {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')

    md_content = ''.join(parts)

    for wh in WECOM_WEBHOOKS:
        if not wh or not wh.strip():
            print('  [企微] 跳过空 Webhook URL')
            continue
        payload = {'msgtype': 'markdown', 'markdown': {'content': md_content}}
        try:
            req = urllib.request.Request(
                wh,
                data=json.dumps(payload).encode('utf-8'),
                headers={'Content-Type': 'application/json'}
            )
            resp = urllib.request.urlopen(req)
            result = json.loads(resp.read())
            if result.get('errcode') == 0:
                print(f'✅ 企微推送成功 → {wh[:60]}...')
            else:
                print(f'❌ 企微推送失败: {result}')
        except Exception as e:
            print(f'❌ 企微推送异常: {e}')


# ============================================================
# 阶段 5: 看板 HTML 生成（v2.0 精装修模板）
# ============================================================

def _fmt_date_str(ts_ms):
    """将毫秒时间戳格式化为 'YYYY-MM-DD HH:MM' 字符串"""
    if not ts_ms:
        return ''
    try:
        dt = datetime.fromtimestamp(int(ts_ms) / 1000)
        return dt.strftime('%Y-%m-%d %H:%M')
    except (OSError, ValueError, TypeError):
        return ''


def _fmt_number(n):
    """格式化为带千分位的字符串"""
    if n is None:
        return '0'
    return f'{int(n):,}'


def _build_record_from_stockout(x):
    """从出库 API 数据构建一条 records 数组（15 个元素）"""
    ts = x.get('applyDate', 0) or 0
    date_str = _fmt_date_str(ts)
    sku_count = int(x.get('skuCount') or 0)
    outed_count = int(x.get('outedCount') or 0)
    unouted_count = int(x.get('unOutedCount') or 0)
    rate = round((outed_count / sku_count * 100) if sku_count > 0 else 0, 1)
    return [
        date_str,                              # [0]  日期字符串
        '出库',                                 # [1]  单据类型
        x.get('outWarehouseName', '') or '',    # [2]  仓库
        x.get('outNo', '') or '',               # [3]  单号
        x.get('relDataId') or '',               # [4]  关联单号
        x.get('brandName', '') or '',           # [5]  品牌
        x.get('goodsName', '') or '',           # [6]  货品名称
        sku_count,                              # [7]  申请数量
        outed_count,                            # [8]  已出库数量
        unouted_count,                          # [9]  未出库数量
        rate,                                   # [10] 完成率
        x.get('field1') or '',                  # [11] 店铺
        x.get('outTypeName', '') or '',         # [12] 业务类型
        int(ts),                                # [13] 时间戳(ms)
        x.get('skuBarcode', '') or '',          # [14] 条码
    ]


def _build_record_from_stockin(x):
    """从入库 API 数据构建一条 records 数组（15 个元素）"""
    ts = x.get('applyDate', 0) or 0
    date_str = _fmt_date_str(ts)
    sku_count = int(x.get('skuCount') or 0)
    inner_count = int(x.get('innerCount') or 0)
    uninner_count = int(x.get('uninnerCount') or 0)
    rate = round((inner_count / sku_count * 100) if sku_count > 0 else 0, 1)
    return [
        date_str,                               # [0]  日期字符串
        '入库',                                  # [1]  单据类型
        x.get('inWarehouseName', '') or '',      # [2]  仓库
        x.get('inNo', '') or '',                 # [3]  单号
        x.get('relDataId') or '',                # [4]  关联单号
        x.get('brandName', '') or '',            # [5]  品牌
        x.get('goodsName', '') or '',            # [6]  货品名称
        sku_count,                               # [7]  申请数量
        inner_count,                             # [8]  已入库数量
        uninner_count,                           # [9]  未入库数量
        rate,                                    # [10] 完成率
        x.get('field1') or '',                   # [11] 店铺
        x.get('inTypeName', '') or '',           # [12] 业务类型
        int(ts),                                 # [13] 时间戳(ms)
        x.get('skuBarcode', '') or '',           # [14] 条码
    ]


def _top_n_by(records, key_fn, value_fn, n=5):
    """按 value_fn 降序排列取 Top N，返回 [(key, count, value), ...]"""
    agg = defaultdict(lambda: {'count': 0, 'value': 0.0})
    for r in records:
        k = key_fn(r)
        agg[k]['count'] += 1
        agg[k]['value'] += value_fn(r)
    sorted_items = sorted(agg.items(), key=lambda kv: kv[1]['value'], reverse=True)[:n]
    return [[k, d['count'], int(d['value'])] for k, d in sorted_items]


def _top_n_simple(records, key_fn, n=10):
    """按 count 降序排列取 Top N，返回 [[key, count], ...]"""
    agg = defaultdict(int)
    for r in records:
        k = key_fn(r)
        agg[k] += 1
    sorted_items = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)[:n]
    return [[k, v] for k, v in sorted_items]


def _build_trend(so_records, si_records):
    """构建月度趋势数据 _trend"""
    from collections import OrderedDict

    # 收集所有月份
    month_set = set()
    for r in so_records:
        ts = r[13]
        if ts:
            try:
                m = datetime.fromtimestamp(ts / 1000).strftime('%Y-%m')
                month_set.add(m)
            except (OSError, ValueError, TypeError):
                pass
    for r in si_records:
        ts = r[13]
        if ts:
            try:
                m = datetime.fromtimestamp(ts / 1000).strftime('%Y-%m')
                month_set.add(m)
            except (OSError, ValueError, TypeError):
                pass

    months = sorted(month_set)
    if not months:
        return None

    # 初始化月度计数
    so_nos = [0] * len(months)
    so_qty = [0] * len(months)
    si_nos = [0] * len(months)
    si_qty = [0] * len(months)
    month_idx = {m: i for i, m in enumerate(months)}

    for r in so_records:
        ts = r[13]
        if ts:
            try:
                m = datetime.fromtimestamp(ts / 1000).strftime('%Y-%m')
                if m in month_idx:
                    idx = month_idx[m]
                    so_nos[idx] += 1
                    so_qty[idx] += r[9]  # unouted count
            except (OSError, ValueError, TypeError):
                pass

    for r in si_records:
        ts = r[13]
        if ts:
            try:
                m = datetime.fromtimestamp(ts / 1000).strftime('%Y-%m')
                if m in month_idx:
                    idx = month_idx[m]
                    si_nos[idx] += 1
                    si_qty[idx] += r[9]  # uninner count
            except (OSError, ValueError, TypeError):
                pass

    return {
        'months': months,
        'so_nos': so_nos,
        'so_qty': so_qty,
        'si_nos': si_nos,
        'si_qty': si_qty,
    }


def generate_dashboard_html(stockout, stockin):
    """
    生成 v2.0 精装修看板 index.html（模板拼接方式）。

    使用 template_before.txt + var DATA = {...} + template_after.txt
    三段拼接生成完整的看板 HTML。
    """
    # ---- 1. 读模板文件 ----
    if not os.path.exists(TEMPLATE_BEFORE_PATH):
        raise FileNotFoundError(f'模板文件不存在: {TEMPLATE_BEFORE_PATH}')
    if not os.path.exists(TEMPLATE_AFTER_PATH):
        raise FileNotFoundError(f'模板文件不存在: {TEMPLATE_AFTER_PATH}')

    with open(TEMPLATE_BEFORE_PATH, 'r', encoding='utf-8') as f:
        template_before = f.read()

    with open(TEMPLATE_AFTER_PATH, 'r', encoding='utf-8') as f:
        template_after = f.read()

    # ---- 2. 构建 records（出库 + 入库统一格式） ----
    so_records = [_build_record_from_stockout(x) for x in stockout]
    si_records = [_build_record_from_stockin(x) for x in stockin]
    all_records = so_records + si_records

    # ---- 3. 计算 KPI 汇总 ----
    so_nos = len(so_records)
    so_total_qty = sum(r[7] for r in so_records)
    so_unouted = sum(r[9] for r in so_records)
    so_rate = round((so_total_qty - so_unouted) / so_total_qty * 100, 1) if so_total_qty > 0 else 0.0

    si_nos = len(si_records)
    si_total_qty = sum(r[7] for r in si_records)
    si_uninner = sum(r[9] for r in si_records)
    si_rate = round((si_total_qty - si_uninner) / si_total_qty * 100, 1) if si_total_qty > 0 else 0.0

    # ---- 4. Top 5 / Top 10 / 分类统计 ----
    # soTop5: 按品牌聚合，按未出库数量降序 Top 5
    so_top5 = _top_n_by(so_records,
                        key_fn=lambda r: r[5],      # brand
                        value_fn=lambda r: r[9],     # unouted count
                        n=5)

    si_top5 = _top_n_by(si_records,
                        key_fn=lambda r: r[5],
                        value_fn=lambda r: r[9],
                        n=5)

    so_brand10 = _top_n_simple(so_records, key_fn=lambda r: r[5], n=10)
    si_brand10 = _top_n_simple(si_records, key_fn=lambda r: r[5], n=10)

    so_types = _top_n_simple(so_records, key_fn=lambda r: r[12], n=50)
    si_types = _top_n_simple(si_records, key_fn=lambda r: r[12], n=50)

    so_wh10 = _top_n_simple(so_records, key_fn=lambda r: r[2], n=10)
    si_wh10 = _top_n_simple(si_records, key_fn=lambda r: r[2], n=10)

    # soComboTop5: 业务类型|仓库 组合，按未出库数量 Top 5
    so_combo_top5 = _top_n_by(so_records,
                              key_fn=lambda r: f'{r[12]}|{r[2]}',
                              value_fn=lambda r: r[9],
                              n=5)

    si_combo_top5 = _top_n_by(si_records,
                              key_fn=lambda r: f'{r[12]}|{r[2]}',
                              value_fn=lambda r: r[9],
                              n=5)

    # ---- 5. 筛选下拉列表（去重排序） ----
    brands = sorted(set(r[5] for r in all_records if r[5]))
    shops = sorted(set(r[11] for r in all_records if r[11]))
    # types 和 whs 在原始 DATA 中都是仓库名称列表（字段名有历史原因）
    whs = sorted(set(r[2] for r in all_records if r[2]))
    types = whs  # types 字段与 whs 内容相同

    # ---- 6. 月度趋势 ----
    trend = _build_trend(so_records, si_records)

    # ---- 7. 构建 DATA JSON ----
    data_obj = {
        'soNos': so_nos,
        'soUnouted': so_unouted,
        'soFmtUnouted': _fmt_number(so_unouted),
        'siNos': si_nos,
        'siUninner': si_uninner,
        'siFmtUninner': _fmt_number(si_uninner),
        'soRate': so_rate,
        'siRate': si_rate,
        'soTop5': so_top5,
        'siTop5': si_top5,
        'soBrand10': so_brand10,
        'siBrand10': si_brand10,
        'soTypes': so_types,
        'siTypes': si_types,
        'soWh10': so_wh10,
        'siWh10': si_wh10,
        'brands': brands,
        'types': types,
        'whs': whs,
        'shops': shops,
        'soComboTop5': so_combo_top5,
        'siComboTop5': si_combo_top5,
        '_trend': trend,
        'records': all_records,
    }

    # 自定义 JSON 序列化：确保整数不变成浮点数，null 保持不变
    json_data = json.dumps(data_obj, ensure_ascii=False, separators=(',', ':'))

    # ---- 8. 拼接并写入 ----
    html = template_before + '\nvar DATA = ' + json_data + ';\n' + template_after

    output_path = os.path.join(SCRIPT_DIR, 'index.html')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(output_path) / 1024
    print(f'✅ 看板 HTML 已生成: {output_path} ({size_kb:.0f} KB)')
    print(f'   出库 {so_nos} 条 + 入库 {si_nos} 条 = 共 {len(all_records)} 条明细')


# ============================================================
# 主流程
# ============================================================

if __name__ == '__main__':
    print('=' * 60)
    print(f'吉客云 B 单未完结监控 · GitHub Actions 版')
    print(f'数据日期: {B_MONITOR_DATE}')
    print(f'开始时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)

    # ---------------------------------
    # 阶段 1: API 数据拉取
    # ---------------------------------
    print('\n[阶段 1/4] API 数据拉取')
    print('-' * 40)

    print('[1.1] 拉取出库申请单...')
    stockout = fetch_stockout()

    print('\n[1.2] 出库 brandName 反查...')
    so_barcodes = set(x.get('skuBarcode', '') for x in stockout if x.get('skuBarcode'))
    print(f'  唯一条码数: {len(so_barcodes)}')
    brand_map = fetch_brand_map(so_barcodes)
    for item in stockout:
        bc = item.get('skuBarcode', '')
        item['brandName'] = brand_map.get(bc, '')
    so_with_brand = sum(1 for x in stockout if x.get('brandName'))
    print(f'  出库 brandName 回填: {so_with_brand}/{len(stockout)}')
    save_json(stockout, 'raw_stockout')

    print('\n[1.3] 拉取入库申请单...')
    stockin = fetch_stockin()
    save_json(stockin, 'raw_stockin')

    print(f'\n✅ 阶段1完成: 出库 {len(stockout)} 单 + 入库 {len(stockin)} 单')

    # ---------------------------------
    # 阶段 2: Excel 生成
    # ---------------------------------
    print(f'\n[阶段 2/5] Excel 报告生成')
    print('-' * 40)
    excel_path = generate_excel(stockout, stockin)

    # ---------------------------------
    # 阶段 3: 邮件发送
    # ---------------------------------
    print(f'\n[阶段 3/5] 邮件发送')
    print('-' * 40)

    if SMTP_USER and SMTP_PASS and TO_ADDR:
        send_email(excel_path, stockout, stockin)
    else:
        print('⚠️ 邮件配置不完整，跳过发送')
        print('  请设置 GitHub Secrets: SMTP_USER, SMTP_PASS, TO_ADDR, CC_ADDR')

    # ---------------------------------
    # 阶段 4: 企微推送
    # ---------------------------------
    print(f'\n[阶段 4/5] 企业微信推送')
    print('-' * 40)

    if WECOM_WEBHOOKS:
        send_wecom(stockout, stockin)
    else:
        print('⚠️ 未配置企微 Webhook，跳过推送')
        print('  请设置 GitHub Secrets: WECOM_WEBHOOK_1, WECOM_WEBHOOK_2')

    # ---------------------------------
    # 阶段 5: 生成看板 HTML
    # ---------------------------------
    print(f'\n[阶段 5/5] 生成看板 HTML')
    print('-' * 40)
    generate_dashboard_html(stockout, stockin)

    # ---------------------------------
    # 完成
    # ---------------------------------
    print('\n' + '=' * 60)
    print(f'✅ 全部完成')
    print(f'完成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'Excel 报告: {excel_path}')
    print('=' * 60)
